# Monkeypatch openpyxl validation error for invalid activePane or state in sheet views
try:
    from openpyxl.worksheet.views import Pane
    from openpyxl.descriptors import String
    
    s_active = String(allow_none=True)
    s_active.name = 'activePane'
    Pane.activePane = s_active
    
    s_state = String(allow_none=True)
    s_state.name = 'state'
    Pane.state = s_state
except Exception:
    pass

import os
import sys
import io
import re
import base64
from datetime import datetime
from flask import Flask, request, jsonify
from flask_cors import CORS
import pandas as pd

# Add current folder to path
sys.path.append(os.path.dirname(__file__))
try:
    from .validator import validate_lazada, validate_shopee, validate_tiktok
except (ImportError, ValueError):
    from validator import validate_lazada, validate_shopee, validate_tiktok

app = Flask(__name__)
CORS(app)

def parse_file(uploaded_file, skip_lazada_rows=False):
    if uploaded_file is None or uploaded_file.filename == '':
        return None
    name = uploaded_file.filename.lower()
    
    if name.endswith('.csv'):
        if skip_lazada_rows:
            df = pd.read_csv(uploaded_file)
            df = df.iloc[3:].reset_index(drop=True)
        else:
            df = pd.read_csv(uploaded_file)
    else:
        if skip_lazada_rows:
            df = pd.read_excel(uploaded_file)
            df = df.iloc[3:].reset_index(drop=True)
        else:
            df = pd.read_excel(uploaded_file)
    return df

def generate_excel_report(lazada_res, shopee_res, tiktok_res):
    from openpyxl.styles import Border, Side, Alignment, PatternFill, Font
    from openpyxl.utils import get_column_letter

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        sheets_written = False
        if lazada_res is not None and not lazada_res.empty:
            lazada_res.to_excel(writer, sheet_name="Lazada Validation", index=False)
            sheets_written = True
        if shopee_res is not None and not shopee_res.empty:
            shopee_res.to_excel(writer, sheet_name="Shopee Validation", index=False)
            sheets_written = True
        if tiktok_res is not None and not tiktok_res.empty:
            tiktok_res.to_excel(writer, sheet_name="TikTok Validation", index=False)
            sheets_written = True
            
        if not sheets_written:
            pd.DataFrame({"Status": ["No validation mismatch data to report."]}).to_excel(writer, sheet_name="Summary", index=False)
            
        # Access worksheets in openpyxl workbook
        workbook = writer.book
        
        # Border style
        thin_side = Side(border_style="thin", color="000000")
        thin_border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)
        
        # Alignment style
        center_align = Alignment(horizontal="center", vertical="center")
        
        # Header styles (Bold and Light Blue highlight background)
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        
        for name in workbook.sheetnames:
            ws = workbook[name]
            
            # Freeze the first row
            ws.freeze_panes = "A2"
            
            # Apply formatting to all cells
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = center_align
            
            # Bold and highlight header cells
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
            
            # Auto-fit column widths
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    val = str(cell.value or '')
                    if len(val) > max_len:
                        max_len = len(val)
                ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
                
    return buffer.getvalue()

@app.route('/api/validate', methods=['POST'])
def validate():
    try:
        # Check files
        if 'all_file' not in request.files or 'tc_inventory' not in request.files:
            return jsonify({'error': 'Missing mandatory reference files (TC All File and TC Inventory)'}), 400
            
        all_file_raw = request.files['all_file']
        tc_inv_raw = request.files['tc_inventory']
        
        lazada_raw = request.files.get('lazada')
        shopee_stock_raw = request.files.get('shopee_stock')
        shopee_status_raw = request.files.get('shopee_status')
        tiktok_active_raw = request.files.get('tiktok_active')
        tiktok_inactive_raw = request.files.get('tiktok_inactive')
        
        # Check config parameters
        country = request.form.get('country', 'SG').upper()
        buffer_type = request.form.get('buffer_type', 'None')
        
        try:
            buffer_val = float(request.form.get('buffer_val', 0))
        except (ValueError, TypeError):
            buffer_val = 0
            
        # Parse Reference Files
        all_df = parse_file(all_file_raw)
        tc_inv_df = parse_file(tc_inv_raw)
        
        if all_df is None or tc_inv_df is None:
            return jsonify({'error': 'Failed to parse mandatory reference files.'}), 400
            
        lazada_results = None
        shopee_sku_results = None
        tiktok_sku_results = None
        
        # Lazada
        if lazada_raw and lazada_raw.filename != '':
            lazada_df = parse_file(lazada_raw, skip_lazada_rows=True)
            lazada_results = validate_lazada(lazada_df, tc_inv_df, all_df, buffer_type, buffer_val)
            
        # Shopee
        if shopee_stock_raw and shopee_stock_raw.filename != '':
            shopee_stock_df = parse_file(shopee_stock_raw)
            shopee_status_df = parse_file(shopee_status_raw) if shopee_status_raw else None
            shopee_sku_results = validate_shopee(shopee_stock_df, shopee_status_df, tc_inv_df, all_df, buffer_type, buffer_val)
            
        # TikTok
        if (tiktok_active_raw and tiktok_active_raw.filename != '') or (tiktok_inactive_raw and tiktok_inactive_raw.filename != ''):
            tiktok_active_df = parse_file(tiktok_active_raw) if tiktok_active_raw else None
            tiktok_inactive_df = parse_file(tiktok_inactive_raw) if tiktok_inactive_raw else None
            tiktok_sku_results = validate_tiktok(tiktok_active_df, tiktok_inactive_df, tc_inv_df, all_df, buffer_type, buffer_val)
            
        # Convert results dataframes to dicts for JSON
        lazada_json = lazada_results.to_dict(orient='records') if lazada_results is not None else None
        shopee_json = shopee_sku_results.to_dict(orient='records') if shopee_sku_results is not None else None
        tiktok_json = tiktok_sku_results.to_dict(orient='records') if tiktok_sku_results is not None else None
        
        # Generate Excel Validation Report
        excel_bytes = generate_excel_report(lazada_results, shopee_sku_results, tiktok_sku_results)
        excel_base64 = base64.b64encode(excel_bytes).decode('utf-8')
        
        # Build filename
        channels_run = []
        if lazada_results is not None:
            channels_run.append("Lazada")
        if shopee_sku_results is not None:
            channels_run.append("Shopee")
        if tiktok_sku_results is not None:
            channels_run.append("TikTok")
            
        channels_str = "_".join(channels_run) if channels_run else "Validation"
        today = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        report_fname = f"DKSH_{channels_str}_Validation_Report_{country}_{today}.xlsx"
        
        return jsonify({
            'lazada': lazada_json,
            'shopee': shopee_json,
            'tiktok': tiktok_json,
            'excel_report': excel_base64,
            'report_filename': report_fname
        })
        
    except Exception as e:
        import traceback
        error_msg = str(e)
        trace = traceback.format_exc()
        print(f"Error: {error_msg}\n{trace}")
        return jsonify({'error': error_msg, 'trace': trace}), 500

if __name__ == '__main__':
    # Run server locally for testing
    app.run(host='0.0.0.0', port=5000, debug=True)
